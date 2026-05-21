from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from updatemenow.exporters.common import REPORT_COLUMNS, item_to_report_row
from updatemenow.models import ScanResult

WORKSHEET_NAME = "Cyber Updates"
TABLE_NAME = "CyberUpdatesTable"
MAX_COLUMN_WIDTH = 80
MIN_COLUMN_WIDTH = 12
HEADER_FILL = PatternFill("solid", fgColor="FF17324D")
HEADER_FONT = Font(bold=True, color="FFFFFFFF")
BODY_BORDER = Border(
    left=Side(style="thin", color="FFD9E2F3"),
    right=Side(style="thin", color="FFD9E2F3"),
    top=Side(style="thin", color="FFD9E2F3"),
    bottom=Side(style="thin", color="FFD9E2F3"),
)
BODY_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
SOURCE_GROUP_FILLS = {
    "government": PatternFill("solid", fgColor="FFF4F8FB"),
    "vulnerability_database": PatternFill("solid", fgColor="FFFDF4F4"),
    "threat_intelligence": PatternFill("solid", fgColor="FFF7F2FD"),
    "security_news": PatternFill("solid", fgColor="FFF4FBF4"),
    "exploit_vulnerability": PatternFill("solid", fgColor="FFFFF8EE"),
}
SOURCE_GROUP_ACCENTS = {
    "government": PatternFill("solid", fgColor="FF4472C4"),
    "vulnerability_database": PatternFill("solid", fgColor="FFC0504D"),
    "threat_intelligence": PatternFill("solid", fgColor="FF8064A2"),
    "security_news": PatternFill("solid", fgColor="FF548235"),
    "exploit_vulnerability": PatternFill("solid", fgColor="FFC99700"),
}
CATEGORY_FILLS = {
    "Exploited Vulnerability": PatternFill("solid", fgColor="FFFCE4D6"),
    "Vulnerability": PatternFill("solid", fgColor="FFFDE9D9"),
    "Vendor Advisory": PatternFill("solid", fgColor="FFEAF2F8"),
    "Government Advisory": PatternFill("solid", fgColor="FFE2F0D9"),
    "Security News": PatternFill("solid", fgColor="FFEDEDED"),
    "Ransomware": PatternFill("solid", fgColor="FFF9D9D9"),
    "Malware": PatternFill("solid", fgColor="FFFCEEDB"),
}


def write_excel_report(result: ScanResult, path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_NAME

    worksheet.append(REPORT_COLUMNS)
    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BODY_BORDER

    reference_time = result.started_at
    for item in result.items:
        worksheet.append(item_to_report_row(item, reference_time))
        row_index = worksheet.max_row
        _style_report_row(worksheet, row_index)
        url_cell = worksheet.cell(row=row_index, column=8)
        if item.url:
            url_cell.hyperlink = item.url
            url_cell.style = "Hyperlink"

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    if worksheet.max_row > 1:
        table = Table(displayName=TABLE_NAME, ref=worksheet.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    _autofit_columns(worksheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def _style_report_row(worksheet, row_index: int) -> None:
    source_group = str(worksheet.cell(row=row_index, column=1).value or "").strip()
    group_fill = SOURCE_GROUP_FILLS.get(source_group)
    accent_fill = SOURCE_GROUP_ACCENTS.get(source_group)
    category = str(worksheet.cell(row=row_index, column=12).value or "").strip()
    category_fill = CATEGORY_FILLS.get(category)

    for column_index in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=row_index, column=column_index)
        cell.border = BODY_BORDER
        cell.alignment = BODY_ALIGNMENT
        if group_fill is not None:
            cell.fill = group_fill

    source_group_cell = worksheet.cell(row=row_index, column=1)
    source_group_cell.fill = accent_fill or source_group_cell.fill
    source_group_cell.font = Font(bold=True, color="FFFFFFFF")

    source_name_cell = worksheet.cell(row=row_index, column=2)
    source_name_cell.font = Font(bold=True, color="FF1F1F1F")

    source_type_cell = worksheet.cell(row=row_index, column=3)
    source_type_cell.alignment = Alignment(horizontal="center", vertical="top")

    age_cell = worksheet.cell(row=row_index, column=5)
    age_cell.font = Font(italic=True, color="FF5A5A5A")
    if str(age_cell.value or "").endswith("m ago"):
        age_cell.fill = PatternFill("solid", fgColor="FFE2F0D9")
    elif str(age_cell.value or "").endswith("h ago"):
        age_cell.fill = PatternFill("solid", fgColor="FFFFF2CC")
    elif str(age_cell.value or "").endswith("d ago"):
        age_cell.fill = PatternFill("solid", fgColor="FFF4CCCC")

    title_cell = worksheet.cell(row=row_index, column=6)
    title_cell.font = Font(bold=True, color="FF1F1F1F")

    description_cell = worksheet.cell(row=row_index, column=7)
    description_cell.alignment = Alignment(wrap_text=True, vertical="top")

    cves_cell = worksheet.cell(row=row_index, column=9)
    cves_cell.font = Font(color="FF7F6000")

    keywords_cell = worksheet.cell(row=row_index, column=10)
    keywords_cell.font = Font(color="FF1F5E92")

    vendors_cell = worksheet.cell(row=row_index, column=11)
    vendors_cell.font = Font(color="FF5B5B5B")

    category_cell = worksheet.cell(row=row_index, column=12)
    if category_fill is not None:
        category_cell.fill = category_fill
    category_cell.font = Font(bold=True, color="FF1F1F1F")
    category_cell.alignment = Alignment(horizontal="center", vertical="top")

    scan_run_time_cell = worksheet.cell(row=row_index, column=13)
    scan_run_time_cell.font = Font(color="FF5A5A5A")


def _autofit_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        width = min(max(max_length + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
