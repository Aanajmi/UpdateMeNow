from datetime import datetime, timedelta, timezone
import json

from openpyxl import load_workbook

from updatemenow.exporters import export_reports
from updatemenow.exporters.common import REPORT_COLUMNS, format_age, report_base_name
from updatemenow.models import (
    CyberUpdateItem,
    ExportFormat,
    ScanRequest,
    ScanResult,
    SourceConfig,
    SourceType,
)

STARTED_AT = datetime(2026, 5, 20, 18, 0, tzinfo=timezone.utc)


def test_export_reports_writes_excel_with_required_sheet_and_columns(tmp_path) -> None:
    request = ScanRequest(hours=6)
    result = _scan_result(items=[_item("item-1")])
    sources = [_source("cisa_kev")]

    exported_result = export_reports(request, result, sources, output_dir=tmp_path)

    export_path = tmp_path / "UpdateMeNow_Report_2026-05-20_6h.xlsx"
    assert exported_result.export_paths == [str(export_path)]
    assert export_path.exists()

    workbook = load_workbook(export_path)
    assert workbook.sheetnames == ["Cyber Updates"]
    worksheet = workbook["Cyber Updates"]
    assert [cell.value for cell in worksheet[1]] == REPORT_COLUMNS
    assert worksheet.freeze_panes == "A2"
    assert worksheet.auto_filter.ref == "A1:M2"
    assert "CyberUpdatesTable" in worksheet.tables
    assert worksheet["A1"].fill.fgColor.rgb == "FF17324D"
    assert worksheet["A1"].font.bold is True
    assert worksheet["A2"].fill.fgColor.rgb == "FF4472C4"
    assert worksheet["F2"].font.bold is True
    assert worksheet["E2"].fill.fgColor.rgb == "FFFFF2CC"
    assert worksheet["L2"].fill.fgColor.rgb == "FFFCE4D6"
    assert worksheet["H2"].hyperlink.target == "https://example.com/advisory"
    assert worksheet["I2"].value == "CVE-2026-12345"
    assert worksheet["J2"].value == "ransomware"
    assert worksheet["K2"].value == "Fortinet"
    assert worksheet.column_dimensions["F"].width > 12


def test_export_reports_writes_json_with_scan_sources_and_items(tmp_path) -> None:
    request = ScanRequest(days=7, exports=[ExportFormat.JSON])
    result = _scan_result(items=[_item("item-1")])
    sources = [_source("cisa_kev")]

    exported_result = export_reports(request, result, sources, output_dir=tmp_path)

    export_path = tmp_path / "UpdateMeNow_Report_2026-05-20_7d.json"
    assert exported_result.export_paths == [str(export_path)]

    payload = json.loads(export_path.read_text(encoding="utf-8"))
    assert payload["scan"]["range"] == "7d"
    assert payload["scan"]["sources_scanned"] == 1
    assert payload["scan"]["raw_items"] == 1
    assert payload["scan"]["final_items"] == 1
    assert payload["scan"]["dedupe_mode"] == "normal"
    assert payload["scan"]["exports"] == ["json"]
    assert payload["scan"]["export_paths"] == [str(export_path)]
    assert payload["sources"][0]["id"] == "cisa_kev"
    assert payload["items"][0]["cves"] == ["CVE-2026-12345"]
    assert payload["items"][0]["keywords_matched"] == ["ransomware"]
    assert payload["items"][0]["vendors_matched"] == ["Fortinet"]


def test_export_reports_writes_excel_and_json_when_requested(tmp_path) -> None:
    request = ScanRequest(exports=[ExportFormat.EXCEL, ExportFormat.JSON])
    result = _scan_result(items=[])

    exported_result = export_reports(request, result, [_source("cisa_kev")], output_dir=tmp_path)

    assert exported_result.export_paths == [
        str(tmp_path / "UpdateMeNow_Report_2026-05-20_6h.xlsx"),
        str(tmp_path / "UpdateMeNow_Report_2026-05-20_6h.json"),
    ]
    assert (tmp_path / "UpdateMeNow_Report_2026-05-20_6h.xlsx").exists()
    assert (tmp_path / "UpdateMeNow_Report_2026-05-20_6h.json").exists()


def test_report_base_name_uses_requested_time_range() -> None:
    result = _scan_result(items=[])

    assert report_base_name(ScanRequest(hours=24), result) == "UpdateMeNow_Report_2026-05-20_24h"
    assert report_base_name(ScanRequest(days=7), result) == "UpdateMeNow_Report_2026-05-20_7d"


def test_format_age_handles_minutes_hours_days_and_unknown() -> None:
    assert format_age(None, STARTED_AT) == "unknown"
    assert format_age(STARTED_AT - timedelta(minutes=30), STARTED_AT) == "30m ago"
    assert format_age(STARTED_AT - timedelta(hours=2), STARTED_AT) == "2h ago"
    assert format_age(STARTED_AT - timedelta(days=1), STARTED_AT) == "1d ago"


def _source(source_id: str) -> SourceConfig:
    return SourceConfig(
        id=source_id,
        name="CISA KEV",
        group="government",
        type=SourceType.API,
        provider="cisa_kev",
        enabled=True,
        default_order=10,
    )


def _scan_result(items: list[CyberUpdateItem]) -> ScanResult:
    return ScanResult(
        scan_id="scan-1",
        started_at=STARTED_AT,
        ended_at=STARTED_AT + timedelta(seconds=2),
        range_hours=6,
        sources_scanned=1,
        raw_item_count=len(items),
        duplicates_removed=0,
        final_item_count=len(items),
        items=items,
        export_paths=[],
    )


def _item(item_id: str) -> CyberUpdateItem:
    return CyberUpdateItem(
        id=item_id,
        source_id="cisa_kev",
        source_name="CISA KEV",
        source_group="government",
        source_type=SourceType.API,
        published_at=STARTED_AT - timedelta(hours=2),
        title="Fortinet advisory for CVE-2026-12345",
        description="Actively exploited ransomware vulnerability.",
        url="https://example.com/advisory",
        cves=["CVE-2026-12345"],
        keywords_matched=["ransomware"],
        vendors_matched=["Fortinet"],
        category="Exploited Vulnerability",
        scan_run_time=STARTED_AT,
    )
